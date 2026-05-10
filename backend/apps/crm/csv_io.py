"""Spreadsheet import (CSV + XLSX, with manual column mapping) and CSV export.

Two-step import flow:

    1. ``preview_import_file(file)`` saves the upload under
       MEDIA_ROOT/imports/<token>.<ext> and returns headers + a sample + a
       suggested mapping. The frontend shows a wizard built from this
       response.
    2. ``apply_import(token, params, user)`` consumes the mapping and
       creates / updates contacts. Errors are collected per-row rather than
       failing the whole batch.

Both CSV and XLSX (.xlsx, modern Office Open XML) are accepted. The legacy
binary .xls format is rejected up front — supporting it would mean adding
``xlrd`` and dealing with a separately deprecated parser. XLSX parsing uses
``openpyxl`` in read-only mode (memory-bounded, lazy iteration).

Export streams the current filtered queryset as a CSV download — no
materialization, so it scales to the SQLite-bound ceiling without OOM.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import logging
import re
import time
import uuid
from pathlib import Path
from typing import Any, Iterable, Iterator

from django.conf import settings
from django.db import transaction
from django.http import StreamingHttpResponse
from rest_framework.exceptions import ValidationError

from .country_codes import normalize_country
from .models import ALLOWED_SOCIAL_KEYS, Contact, ContactLabel


logger = logging.getLogger(__name__)


# Where pending uploads live between preview and apply. Cleaned up by the
# apply step; orphans get garbage-collected by ``_purge_stale_imports`` on
# every preview call (cheap O(n) scan of the directory).
_IMPORT_DIRNAME = "imports"
_IMPORT_TTL_SECONDS = 60 * 60  # 1h: enough for a user to complete the wizard
_MAX_BYTES = 20 * 1024 * 1024  # 20 MB cap on upload size
_MAX_PREVIEW_ROWS = 20

# Hard cap on how many rows we'll process synchronously. Above this we ask
# the user to split the file — Phase 1 has no background worker.
MAX_IMPORT_ROWS = 10_000

# File-kind enum (lower-case strings to keep them serializable and obvious
# in logs / API responses).
KIND_CSV = "csv"
KIND_XLSX = "xlsx"
_KIND_EXTENSIONS = {KIND_CSV: ".csv", KIND_XLSX: ".xlsx"}


# Synonyms used to auto-suggest a mapping from a CSV header. Matched on the
# normalized header (lowercase, alphanumeric + spaces, collapsed whitespace).
_HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "first_name": ("first name", "firstname", "fname", "given name", "first", "prenom", "prénom"),
    "last_name": ("last name", "lastname", "lname", "surname", "family name", "last", "nom"),
    "company": ("company", "company name", "organization", "organisation", "org", "business", "employer", "entreprise", "société", "societe"),
    "industry": ("industry", "sector", "vertical", "company type", "business type", "secteur", "industrie"),
    "job_title": ("job title", "title", "role", "position", "job role", "function", "designation", "fonction", "poste"),
    "email": ("email", "email address", "e mail", "mail", "courriel"),
    "phone": ("phone", "phone number", "phonenumber", "tel", "telephone", "mobile", "cell", "cellphone", "téléphone"),
    "address_line1": ("address", "address line 1", "address1", "street", "street address", "addr1", "line 1", "adresse"),
    "address_line2": ("address line 2", "address2", "addr2", "line 2", "apt", "apartment", "unit", "suite"),
    "city": ("city", "town", "ville"),
    "region": ("region", "state", "province", "county", "département", "departement"),
    "postal_code": ("postal code", "postcode", "post code", "zip", "zipcode", "zip code", "code postal"),
    "country": ("country", "nation", "pays"),
    "websites": ("website", "websites", "url", "web", "site", "homepage"),
    "socials.linkedin": ("linkedin", "linkedin url", "linked in"),
    "socials.twitter": ("twitter", "twitter url", "x", "x url", "x handle", "twitter handle"),
    "socials.facebook": ("facebook", "facebook url", "fb", "fb url"),
    "socials.instagram": ("instagram", "instagram url", "ig", "ig handle", "insta"),
    "labels": ("labels", "tags", "categories", "category", "status", "type", "segment"),
    "notes": ("notes", "comment", "comments", "description", "memo", "remarks", "remarque"),
}


# Targets the apply step understands. Anything else in the mapping value is
# treated as ``[ignore]`` for safety.
_VALID_TARGETS: set[str] = {
    "company", "first_name", "last_name",
    "industry", "job_title",
    "email", "phone",
    "address_line1", "address_line2", "city", "region", "postal_code", "country",
    "websites", "labels", "notes",
    *(f"socials.{k}" for k in ALLOWED_SOCIAL_KEYS),
    "[ignore]",
}


# ── Filesystem helpers ───────────────────────────────────────────────────


def _import_dir() -> Path:
    p = Path(settings.MEDIA_ROOT) / _IMPORT_DIRNAME
    p.mkdir(parents=True, exist_ok=True)
    return p


def _purge_stale_imports() -> None:
    cutoff = time.time() - _IMPORT_TTL_SECONDS
    for entry in _import_dir().iterdir():
        try:
            if not entry.is_file():
                continue
            if entry.suffix not in _KIND_EXTENSIONS.values():
                continue
            if entry.stat().st_mtime < cutoff:
                entry.unlink(missing_ok=True)
        except OSError:  # pragma: no cover
            continue


def _detect_kind(*, filename: str | None, raw: bytes) -> str:
    """Determine whether the upload is CSV or XLSX, rejecting old .xls."""
    # XLSX is a ZIP archive — magic bytes 'PK\x03\x04'.
    if raw[:4] == b"PK\x03\x04":
        return KIND_XLSX
    # OLE2 compound file (legacy .xls). We don't support it.
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise ValidationError(
            {
                "file": (
                    "Old .xls (Excel 97–2003) files aren't supported. "
                    "Save as .xlsx or export to CSV."
                )
            }
        )
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return KIND_XLSX
    if name.endswith(".xls"):
        raise ValidationError(
            {
                "file": (
                    "Old .xls (Excel 97–2003) files aren't supported. "
                    "Save as .xlsx or export to CSV."
                )
            }
        )
    return KIND_CSV


def _resolve_token_path(token: str) -> tuple[Path, str]:
    """Find the on-disk file for ``token`` and return (path, kind).

    Tokens carry no kind hint; we store with the original extension and
    probe both possibilities here.
    """
    safe = Path(token).name  # strip path traversal attempts
    base = _import_dir() / safe
    for kind, ext in _KIND_EXTENSIONS.items():
        candidate = base.with_suffix(ext)
        if candidate.exists():
            return candidate, kind
    raise ValidationError({"token": "Import token not found or expired."})


# ── Decoding & header normalization ──────────────────────────────────────


def _decode_csv_bytes(raw: bytes) -> str:
    """Try a few common encodings — Excel exports are usually utf-8-sig or cp1252."""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _sniff_dialect(text: str) -> csv.Dialect:
    """Detect delimiter (comma / semicolon / tab). Falls back to comma."""
    try:
        return csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        return csv.excel  # comma, default


def _cell_to_str(value: Any) -> str:
    """XLSX cells come back as native types; coerce to a stripped string.

    Datetimes become ISO strings (so the importer's downstream logic can
    handle them as text the same way a CSV cell would). Booleans become
    ``"true"``/``"false"`` lower-case for stable round-trips.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # Excel stores ints as floats. ``42.0`` should look like ``42`` in
        # the CSV-equivalent of the row, otherwise the email-as-int trick
        # ("123 Main St" parsed back as a number) yields ``123.0``.
        return str(int(value))
    return str(value).strip()


_NORM_RE = re.compile(r"[^a-z0-9]+")


def _normalize_header(s: str) -> str:
    return _NORM_RE.sub(" ", s.strip().lower()).strip()


def _suggest_mapping(headers: Iterable[str]) -> dict[str, str]:
    """Match each source header to a target field via header-synonym table."""
    # Pre-normalize the synonyms once so ``startswith`` / equality runs on
    # already-cleaned strings.
    norm_synonyms: dict[str, str] = {}
    for target, syns in _HEADER_SYNONYMS.items():
        for s in syns:
            norm_synonyms[_normalize_header(s)] = target

    out: dict[str, str] = {}
    for h in headers:
        n = _normalize_header(h)
        if n in norm_synonyms:
            out[h] = norm_synonyms[n]
        else:
            out[h] = "[ignore]"
    return out


# ── Format-specific readers ──────────────────────────────────────────────


def _read_csv_table(raw: bytes) -> tuple[list[str], list[list[str]], int, str]:
    """Parse CSV bytes into ``(headers, sample, total_data_rows, delimiter)``."""
    text = _decode_csv_bytes(raw)
    dialect = _sniff_dialect(text)
    rows = list(csv.reader(io.StringIO(text), dialect=dialect))
    if not rows:
        raise ValidationError({"file": "No rows found."})
    headers = [str(h) for h in rows[0]]
    if not any(h.strip() for h in headers):
        raise ValidationError({"file": "No header row found."})
    sample = [list(r) for r in rows[1 : 1 + _MAX_PREVIEW_ROWS]]
    row_count = max(0, len(rows) - 1)
    return headers, sample, row_count, dialect.delimiter


def _read_xlsx_table(path: Path) -> tuple[list[str], list[list[str]], int, str]:
    """Parse the active sheet of an ``.xlsx`` file via openpyxl read-only mode."""
    # Imported lazily so the rest of the module loads cleanly even if the
    # dependency is missing in some weird environment.
    from openpyxl import load_workbook  # noqa: WPS433 (intentional local import)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValidationError(
            {"file": f"Could not open .xlsx file: {exc}"}
        ) from exc

    try:
        ws = wb.active
        if ws is None:
            raise ValidationError({"file": "Workbook has no active sheet."})

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_tuple = next(rows_iter)
        except StopIteration as exc:
            raise ValidationError({"file": "Empty workbook."}) from exc

        headers = [_cell_to_str(c) for c in header_tuple]
        if not any(h for h in headers):
            raise ValidationError({"file": "No header row found."})

        sample: list[list[str]] = []
        row_count = 0
        for row in rows_iter:
            cells = [_cell_to_str(c) for c in row]
            if not any(cells):
                # Wholly empty row — skip without counting (XLSX often has
                # phantom trailing empties from formatted-but-empty cells).
                continue
            if len(sample) < _MAX_PREVIEW_ROWS:
                sample.append(cells)
            row_count += 1
            if row_count > MAX_IMPORT_ROWS:
                # Bail without finishing the iterator — saves time on huge
                # files we'd reject anyway.
                row_count = MAX_IMPORT_ROWS + 1
                break

        return headers, sample, row_count, "xlsx"
    finally:
        wb.close()


# ── Public API: preview ──────────────────────────────────────────────────


def preview_import_file(uploaded) -> dict[str, Any]:
    """Save the upload, parse headers + a small sample, and suggest a mapping.

    Accepts CSV and XLSX. Returns a dict with the format under ``"format"``
    so the frontend can hint at how the file was parsed (delimiter for CSV,
    sheet for XLSX, etc.).
    """
    _purge_stale_imports()

    raw = uploaded.read()
    if not raw:
        raise ValidationError({"file": "File is empty."})
    if len(raw) > _MAX_BYTES:
        raise ValidationError(
            {"file": f"Upload too large ({len(raw)} bytes > {_MAX_BYTES})."}
        )

    kind = _detect_kind(filename=getattr(uploaded, "name", None), raw=raw)

    token = uuid.uuid4().hex
    dest = _import_dir() / f"{token}{_KIND_EXTENSIONS[kind]}"
    dest.write_bytes(raw)

    try:
        if kind == KIND_XLSX:
            headers, sample, row_count, delimiter = _read_xlsx_table(dest)
        else:
            headers, sample, row_count, delimiter = _read_csv_table(raw)
    except ValidationError:
        dest.unlink(missing_ok=True)
        raise

    if row_count > MAX_IMPORT_ROWS:
        dest.unlink(missing_ok=True)
        raise ValidationError(
            {
                "file": (
                    f"File has {row_count} rows; the synchronous importer "
                    f"caps at {MAX_IMPORT_ROWS}. Split the file and retry."
                )
            }
        )

    return {
        "token": token,
        "format": kind,
        "headers": headers,
        "sample_rows": sample,
        "row_count": row_count,
        "delimiter": delimiter,
        "suggested_mapping": _suggest_mapping(headers),
        "valid_targets": sorted(_VALID_TARGETS),
    }


# Backwards-compat alias for callers that still want the old name.
preview_csv = preview_import_file


# ── Public API: apply ────────────────────────────────────────────────────


def _iter_csv_dict_rows(path: Path) -> Iterator[dict[str, str]]:
    raw = path.read_bytes()
    text = _decode_csv_bytes(raw)
    dialect = _sniff_dialect(text)
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    for row in reader:
        yield {
            (k or ""): ("" if v is None else str(v))
            for k, v in row.items()
            if k is not None
        }


def _iter_xlsx_dict_rows(path: Path) -> Iterator[dict[str, str]]:
    """Yield each data row of an XLSX as a header→stringified-cell dict."""
    from openpyxl import load_workbook  # noqa: WPS433

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_tuple = next(rows_iter)
        except StopIteration:
            return
        headers = [_cell_to_str(c) for c in header_tuple]
        for row in rows_iter:
            cells = [_cell_to_str(c) for c in row]
            if not any(cells):
                # Pass through as empty dict so the row counter still advances
                # — matches CSV behaviour where DictReader yields blanks.
                yield {h: "" for h in headers}
                continue
            yield {
                h: (cells[i] if i < len(cells) else "")
                for i, h in enumerate(headers)
            }
    finally:
        wb.close()


def apply_import(
    token: str,
    *,
    mapping: dict[str, str],
    dedupe: str = "email",
    on_conflict: str = "skip",
    user=None,
) -> dict[str, Any]:
    file_path, kind = _resolve_token_path(token)

    # Sanitise mapping: drop unknown targets early.
    mapping = {
        src: (tgt if tgt in _VALID_TARGETS else "[ignore]")
        for src, tgt in mapping.items()
    }

    if kind == KIND_XLSX:
        row_iter = _iter_xlsx_dict_rows(file_path)
    else:
        row_iter = _iter_csv_dict_rows(file_path)

    label_cache: dict[str, ContactLabel] = {}
    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    auth_user = user if user is not None and getattr(user, "is_authenticated", False) else None

    # Row numbering: +1 for header, +1 for 1-indexing → first data row is 2.
    for row_num, row in enumerate(row_iter, start=2):
        try:
            data, label_names = _row_to_contact_data(row, mapping)

            if not _has_meaningful_data(data, label_names):
                skipped += 1
                continue

            existing = _find_existing(data, dedupe) if dedupe != "none" else None

            if existing is not None:
                if on_conflict == "skip":
                    skipped += 1
                    continue
                _merge_into(existing, data)
                with transaction.atomic():
                    existing.save()
                    if label_names:
                        _attach_labels(existing, label_names, label_cache)
                updated += 1
            else:
                with transaction.atomic():
                    contact = Contact(created_by=auth_user, **data)
                    contact.save()
                    if label_names:
                        _attach_labels(contact, label_names, label_cache)
                created += 1
        except Exception as exc:  # noqa: BLE001 — we WANT to keep going
            errors.append({"row": row_num, "reason": _format_error(exc)})
            if len(errors) > 200:
                errors.append({"row": row_num, "reason": "too many errors; aborting"})
                break

    file_path.unlink(missing_ok=True)

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


# Backwards-compat alias.
apply_csv_import = apply_import


# ── Row → data dict ──────────────────────────────────────────────────────


def _row_to_contact_data(
    row: dict[str, Any],
    mapping: dict[str, str],
) -> tuple[dict[str, Any], list[str]]:
    data: dict[str, Any] = {
        "company": "",
        "first_name": "",
        "last_name": "",
        "industry": "",
        "job_title": "",
        "email": "",
        "phone": "",
        "address_line1": "",
        "address_line2": "",
        "city": "",
        "region": "",
        "postal_code": "",
        "country": "",
        "websites": [],
        "socials": {},
        "notes": "",
    }
    label_names: list[str] = []

    for source_col, raw_value in row.items():
        if source_col is None:
            continue
        target = mapping.get(source_col, "[ignore]")
        if target == "[ignore]":
            continue
        if raw_value is None:
            continue
        v = str(raw_value).strip()
        if not v:
            continue

        if target == "country":
            normalized = normalize_country(v)
            data["country"] = normalized or v[:2].upper()
        elif target == "websites":
            for u in (s.strip() for s in v.split(",")):
                if u and u not in data["websites"]:
                    data["websites"].append(u)
        elif target == "labels":
            for ln in (s.strip() for s in v.split(",")):
                if ln and ln not in label_names:
                    label_names.append(ln)
        elif target.startswith("socials."):
            key = target.split(".", 1)[1]
            if key in ALLOWED_SOCIAL_KEYS:
                data["socials"][key] = v
        elif target in data:
            # Coerce email lower in storage; everything else stays as-typed.
            if target == "email":
                data[target] = v.lower()
            else:
                data[target] = v

    return data, label_names


def _has_meaningful_data(data: dict[str, Any], labels: list[str]) -> bool:
    """Return True if at least one substantive field has a value."""
    for key in (
        "company",
        "first_name",
        "last_name",
        "industry",
        "job_title",
        "email",
        "phone",
        "city",
        "country",
        "notes",
    ):
        if data.get(key):
            return True
    if data.get("websites"):
        return True
    if data.get("socials"):
        return True
    if labels:
        return True
    return False


# ── Dedupe ───────────────────────────────────────────────────────────────


def _find_existing(data: dict[str, Any], strategy: str) -> Contact | None:
    if strategy == "email":
        if email := data.get("email"):
            return Contact.objects.filter(email__iexact=email).first()
        return None
    if strategy == "name+company":
        fn = data.get("first_name") or ""
        ln = data.get("last_name") or ""
        co = data.get("company") or ""
        if fn and ln and co:
            return Contact.objects.filter(
                first_name__iexact=fn,
                last_name__iexact=ln,
                company__iexact=co,
            ).first()
        return None
    return None


def _merge_into(existing: Contact, data: dict[str, Any]) -> None:
    """Fill blanks on ``existing`` with non-blank values from ``data``.

    Re-importing the same export should not nuke better data we already had
    — values already on the record win.
    """
    for key in (
        "company",
        "first_name",
        "last_name",
        "industry",
        "job_title",
        "email",
        "phone",
        "address_line1",
        "address_line2",
        "city",
        "region",
        "postal_code",
        "country",
        "notes",
    ):
        new_value = data.get(key) or ""
        if new_value and not getattr(existing, key, ""):
            setattr(existing, key, new_value)
    # Lists: union, preserving existing order.
    if data.get("websites"):
        merged = list(existing.websites or [])
        for u in data["websites"]:
            if u not in merged:
                merged.append(u)
        existing.websites = merged
    if data.get("socials"):
        merged_s = dict(existing.socials or {})
        for k, v in data["socials"].items():
            merged_s.setdefault(k, v)
        existing.socials = merged_s


def _attach_labels(
    contact: Contact,
    names: list[str],
    cache: dict[str, ContactLabel],
) -> None:
    for name in names:
        cached = cache.get(name.lower())
        if cached is None:
            cached, _ = ContactLabel.objects.get_or_create(name=name)
            cache[name.lower()] = cached
        contact.labels.add(cached)


def _format_error(exc: Exception) -> str:
    msg = str(exc)
    if not msg:
        msg = exc.__class__.__name__
    if len(msg) > 200:
        msg = msg[:200] + "…"
    return msg


# ── Export ───────────────────────────────────────────────────────────────


_EXPORT_HEADERS = [
    "key",
    "company",
    "first_name",
    "last_name",
    "industry",
    "job_title",
    "email",
    "phone",
    "address_line1",
    "address_line2",
    "city",
    "region",
    "postal_code",
    "country",
    "websites",
    "linkedin",
    "twitter",
    "facebook",
    "instagram",
    "labels",
    "notes",
    "created_at",
    "updated_at",
]


class _Echo:
    """File-like object whose write() returns whatever it was given.

    Lets ``csv.writer`` serialize each row into the streaming response
    without ever buffering more than one row.
    """

    def write(self, value):  # noqa: D401
        return value


def stream_csv_export(qs) -> StreamingHttpResponse:
    """Stream the filtered queryset as a CSV download."""
    # Force the prefetch + select_related already on the queryset; iterator()
    # disables the second-level prefetch, so we use chunk_size + manual loop.

    pseudo = _Echo()
    writer = csv.writer(pseudo)

    def gen():
        yield writer.writerow(_EXPORT_HEADERS)
        for c in qs.iterator(chunk_size=500):
            socials = c.socials or {}
            yield writer.writerow(
                [
                    c.key,
                    c.company,
                    c.first_name,
                    c.last_name,
                    c.industry,
                    c.job_title,
                    c.email,
                    c.phone,
                    c.address_line1,
                    c.address_line2,
                    c.city,
                    c.region,
                    c.postal_code,
                    c.country,
                    ", ".join(c.websites or []),
                    socials.get("linkedin", ""),
                    socials.get("twitter", ""),
                    socials.get("facebook", ""),
                    socials.get("instagram", ""),
                    ", ".join(label.name for label in c.labels.all()),
                    c.notes,
                    c.created_at.isoformat() if c.created_at else "",
                    c.updated_at.isoformat() if c.updated_at else "",
                ]
            )

    response = StreamingHttpResponse(gen(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="contacts.csv"'
    return response
